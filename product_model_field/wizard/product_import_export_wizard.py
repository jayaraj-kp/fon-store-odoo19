import base64
import io
import logging

from odoo import models, fields, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


# ─── Column mapping: (header label, field on product.template) ──────────────
EXPORT_COLUMNS = [
    ('Internal Reference',  'default_code'),
    ('Product Name',        'name'),
    ('Model',               'model_number'),
    ('Sales Price',         'list_price'),
    ('Cost',                'standard_price'),
    ('MRP Price',           'mrp_price'),
    ('Category',            'categ_id'),
    ('Barcode',             'barcode'),
]

IMPORT_REQUIRED = {'Product Name'}   # columns that must be non-empty


class ProductImportExportWizard(models.TransientModel):
    _name = 'product.import.export.wizard'
    _description = 'Product Import / Export Wizard (with Model & MRP Price)'

    # ── mode ────────────────────────────────────────────────────────────────
    mode = fields.Selection(
        [('export', 'Export to Excel'), ('import', 'Import from Excel')],
        string='Action', default='export', required=True,
    )

    # ── export ───────────────────────────────────────────────────────────────
    export_file      = fields.Binary(string='Download File',   readonly=True)
    export_filename  = fields.Char(string='File Name',         readonly=True)

    # ── import ───────────────────────────────────────────────────────────────
    import_file      = fields.Binary(string='Upload Excel File')
    import_filename  = fields.Char(string='File Name')
    import_result    = fields.Text(string='Import Result', readonly=True)

    # ────────────────────────────────────────────────────────────────────────
    #  EXPORT
    # ────────────────────────────────────────────────────────────────────────
    def action_export(self):
        if openpyxl is None:
            raise UserError(_('openpyxl library is not installed on this server.'))

        products = self.env['product.template'].search([], order='name')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Products'

        # ── styles ──────────────────────────────────────────────────────────
        header_fill  = PatternFill('solid', fgColor='4F46E5')
        header_font  = Font(bold=True, color='FFFFFF', size=11)
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin         = Side(style='thin', color='D1D5DB')
        border       = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── header row ───────────────────────────────────────────────────────
        headers = [col[0] for col in EXPORT_COLUMNS]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = header_align
            cell.border = border

        ws.row_dimensions[1].height = 30

        # ── data rows ────────────────────────────────────────────────────────
        number_cols = {'Sales Price', 'Cost', 'MRP Price'}
        for row_idx, product in enumerate(products, 2):
            for col_idx, (header, fname) in enumerate(EXPORT_COLUMNS, 1):
                val = getattr(product, fname, '')
                # Many2one → display name
                if hasattr(val, 'name'):
                    val = val.name or ''
                elif val is False:
                    val = ''

                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
                if header in number_cols:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        # ── column widths ────────────────────────────────────────────────────
        col_widths = [20, 40, 18, 14, 14, 14, 25, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = 'A2'

        # ── save ─────────────────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        self.write({
            'export_file':     base64.b64encode(buf.read()),
            'export_filename': 'products_export.xlsx',
        })

        return {
            'type':      'ir.actions.act_window',
            'res_model': self._name,
            'res_id':    self.id,
            'view_mode': 'form',
            'target':    'new',
        }

    # ────────────────────────────────────────────────────────────────────────
    #  DOWNLOAD TEMPLATE
    # ────────────────────────────────────────────────────────────────────────
    def action_download_template(self):
        if openpyxl is None:
            raise UserError(_('openpyxl library is not installed on this server.'))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Products Import'

        header_fill  = PatternFill('solid', fgColor='4F46E5')
        header_font  = Font(bold=True, color='FFFFFF', size=11)
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin         = Side(style='thin', color='D1D5DB')
        border       = Border(left=thin, right=thin, top=thin, bottom=thin)

        required_fill = PatternFill('solid', fgColor='FEF3C7')

        headers = [col[0] for col in EXPORT_COLUMNS]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = header_align
            cell.border    = border
            if header in IMPORT_REQUIRED:
                cell.fill = required_fill
                cell.font = Font(bold=True, color='92400E', size=11)

        # sample row
        sample = ['REF001', 'Sample Product', 'IP14PM', 749.00, 239.00, 999.00, 'All', '']
        for col_idx, val in enumerate(sample, 1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.border    = border
            cell.alignment = Alignment(horizontal='left')

        col_widths = [20, 40, 18, 14, 14, 14, 25, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        self.write({
            'export_file':     base64.b64encode(buf.read()),
            'export_filename': 'products_import_template.xlsx',
        })

        return {
            'type':      'ir.actions.act_window',
            'res_model': self._name,
            'res_id':    self.id,
            'view_mode': 'form',
            'target':    'new',
        }

    # ────────────────────────────────────────────────────────────────────────
    #  IMPORT
    # ────────────────────────────────────────────────────────────────────────
    def action_import(self):
        if openpyxl is None:
            raise UserError(_('openpyxl library is not installed on this server.'))
        if not self.import_file:
            raise UserError(_('Please upload an Excel file first.'))

        raw  = base64.b64decode(self.import_file)
        buf  = io.BytesIO(raw)
        wb   = openpyxl.load_workbook(buf, data_only=True)
        ws   = wb.active

        # build header map  {label: col_index}
        header_row  = [cell.value for cell in ws[1]]
        header_map  = {str(v).strip(): i for i, v in enumerate(header_row) if v}

        # validate that at minimum "Product Name" exists
        missing = IMPORT_REQUIRED - set(header_map.keys())
        if missing:
            raise UserError(_(
                'Required column(s) missing in the uploaded file: %s'
            ) % ', '.join(missing))

        ProductTpl = self.env['product.template']

        created = updated = errors = 0
        error_lines = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            def val(header):
                idx = header_map.get(header)
                return row[idx] if idx is not None and idx < len(row) else None

            name = val('Product Name')
            if not name:
                continue        # skip blank rows

            try:
                vals = {'name': str(name).strip()}

                model_num = val('Model')
                if model_num is not None and str(model_num).strip():
                    vals['model_number'] = str(model_num).strip()

                mrp = val('MRP Price')
                if mrp is not None and mrp != '':
                    try:
                        vals['mrp_price'] = float(mrp)
                    except (ValueError, TypeError):
                        error_lines.append(f'Row {row_num}: Invalid MRP Price "{mrp}"')
                        errors += 1
                        continue

                sales_price = val('Sales Price')
                if sales_price is not None and sales_price != '':
                    try:
                        vals['list_price'] = float(sales_price)
                    except (ValueError, TypeError):
                        pass

                cost = val('Cost')
                if cost is not None and cost != '':
                    try:
                        vals['standard_price'] = float(cost)
                    except (ValueError, TypeError):
                        pass

                ref = val('Internal Reference')
                if ref is not None and str(ref).strip():
                    vals['default_code'] = str(ref).strip()

                barcode = val('Barcode')
                if barcode is not None and str(barcode).strip():
                    vals['barcode'] = str(barcode).strip()

                categ_name = val('Category')
                if categ_name and str(categ_name).strip():
                    categ = self.env['product.category'].search(
                        [('name', '=', str(categ_name).strip())], limit=1
                    )
                    if categ:
                        vals['categ_id'] = categ.id

                # match by Internal Reference first, then by name
                existing = False
                if vals.get('default_code'):
                    existing = ProductTpl.search(
                        [('default_code', '=', vals['default_code'])], limit=1
                    )
                if not existing:
                    existing = ProductTpl.search(
                        [('name', '=', vals['name'])], limit=1
                    )

                if existing:
                    existing.write(vals)
                    updated += 1
                else:
                    ProductTpl.create(vals)
                    created += 1

            except Exception as e:
                errors += 1
                error_lines.append(f'Row {row_num}: {e}')
                _logger.exception('Product import error at row %s', row_num)

        result = (
            f'✅ Import complete!\n'
            f'   • Created : {created}\n'
            f'   • Updated : {updated}\n'
            f'   • Errors  : {errors}\n'
        )
        if error_lines:
            result += '\nError details:\n' + '\n'.join(error_lines)

        self.import_result = result

        return {
            'type':      'ir.actions.act_window',
            'res_model': self._name,
            'res_id':    self.id,
            'view_mode': 'form',
            'target':    'new',
        }
